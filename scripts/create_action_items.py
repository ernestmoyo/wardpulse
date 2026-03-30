"""Generate action items Excel spreadsheet for WardPulse — Google Sheets compatible."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
hfont = Font(bold=True, color="FFFFFF", size=11)
fills = {
    "ernest": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "asher": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
    "joint": PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
    "decisions": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
    "risks": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    "HIGH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "LOW": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
wrap = Alignment(wrap_text=True, vertical="top")


def header(ws, row, key):
    for cell in ws[row]:
        cell.font = hfont
        cell.fill = fills[key]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def row_style(ws, row, prio_col=2):
    for cell in ws[row]:
        cell.alignment = wrap
        cell.border = thin
    v = str(ws.cell(row, prio_col).value or "").upper()
    for lvl in ("HIGH", "MEDIUM", "LOW"):
        if lvl in v:
            ws.cell(row, prio_col).fill = fills[lvl]


# ===== Ernest =====
ws1 = wb.active
ws1.title = "Ernest Action Items"
ws1.append(["#", "Priority", "Action Item", "Due Date", "Status", "Notes"])
header(ws1, 1, "ernest")
for i, item in enumerate([
    ["HIGH", "Draft and send MOU to Asher", "2026-03-30", "DONE", "MOU created as Word doc"],
    ["HIGH", "Create data input template for Asher", "2026-04-01", "PENDING", "Template for real ward data entry"],
    ["HIGH", "Write ministry data request list", "2026-04-01", "DONE", "Includes template letter"],
    ["MEDIUM", "Replace simulated data with real data", "After Asher shares", "BLOCKED", "Waiting on Google Forms link"],
    ["MEDIUM", "Set up GitHub repo access", "2026-03-30", "DONE", "Add Asher as collaborator"],
    ["LOW", "Research domain purchase (wardpulse.org)", "2026-04-07", "PENDING", ""],
    ["LOW", "Prepare launch countdown visuals", "2026-04-14", "PENDING", "10 days to go format"],
], start=1):
    ws1.append([i] + item)
    row_style(ws1, ws1.max_row)

# ===== Asher =====
ws2 = wb.create_sheet("Asher Action Items")
ws2.append(["#", "Priority", "Action Item", "Due Date", "Status", "Notes"])
header(ws2, 1, "asher")
for i, item in enumerate([
    ["HIGH", "Share Google Forms link with ward data", "2026-04-01", "PENDING", "Has data from German-Zimbabwean colleague"],
    ["HIGH", "Research for-profit registration in Zimbabwe", "2026-04-07", "PENDING", "For-profit due to PVO Act"],
    ["HIGH", "Talk to business partner about registering", "2026-04-07", "PENDING", "Printing company partner provides office"],
    ["MEDIUM", "Request shapefiles from City of Harare", "2026-04-07", "PENDING", "Water, sewer, roads, lighting"],
    ["MEDIUM", "Photograph physical maps at Council office", "2026-04-07", "PENDING", "Ernest will digitize"],
    ["MEDIUM", "Review and sign MOU", "2026-04-02", "PENDING", "Sent via email"],
    ["LOW", "Share with YALI/US alumni network", "TBD", "PENDING", "After launch"],
], start=1):
    ws2.append([i] + item)
    row_style(ws2, ws2.max_row)

# ===== Joint =====
ws3 = wb.create_sheet("Joint Action Items")
ws3.append(["#", "Priority", "Action Item", "Due Date", "Owner", "Status", "Notes"])
header(ws3, 1, "joint")
for i, item in enumerate([
    ["HIGH", "Review and sign MOU", "2026-04-02", "Both", "PENDING", "Ernest drafts, Asher reviews"],
    ["MEDIUM", "Schedule next meeting", "Week of 2026-04-06", "Both", "PENDING", ""],
    ["MEDIUM", "Create WardPulse social media page", "2026-04-07", "Both", "PENDING", "LinkedIn"],
    ["MEDIUM", "Plan countdown launch campaign", "2026-04-14", "Both", "PENDING", ""],
    ["LOW", "Discuss monetization strategy", "Next meeting", "Both", "PENDING", "Training, consulting, grants"],
], start=1):
    ws3.append([i] + item)
    row_style(ws3, ws3.max_row)

# ===== Key Decisions =====
ws4 = wb.create_sheet("Key Decisions")
ws4.append(["Decision", "Detail", "Rationale"])
header(ws4, 1, "decisions")
for item in [
    ["Entity type", "For-profit company", "PVO Act makes non-profit registration difficult"],
    ["Ownership", "50/50 split", "Equal partnership agreed by both parties"],
    ["Scope", "Harare(46) + Chitungwiza(25) + Epworth(7) = 78 wards", "Start local, expand later"],
    ["Open source", "Platform code open source", "Transparency and credibility"],
    ["Launch", "Countdown campaign (10 days to go)", "Leverage both networks"],
    ["Legal basis", "Asher operates under ZMC accreditation", "Right to gather civic data"],
    ["Tech stack", "Next.js + MapLibre + JSON (Supabase later)", "Fast dev, no cloud dependency"],
]:
    ws4.append(item)
    for cell in ws4[ws4.max_row]:
        cell.alignment = wrap
        cell.border = thin

# ===== Risks =====
ws5 = wb.create_sheet("Risks")
ws5.append(["Risk", "Severity", "Mitigation"])
header(ws5, 1, "risks")
for item in [
    ["Asher self-funded, no revenue yet", "HIGH", "Plan monetization early; seek grants via YALI"],
    ["PVO Act / Cyber Act regulatory risk", "MEDIUM", "Register as for-profit; frame as journalism"],
    ["No immediate revenue stream", "MEDIUM", "Training, consulting, municipal budget inclusion"],
    ["Data accuracy depends on citizen reports", "MEDIUM", "Seed with Asher-verified data; admin review"],
    ["Government pushback on accountability data", "LOW", "Minister and Mayor already aware/supportive"],
    ["Technical scalability (JSON files)", "LOW", "Migrate to Supabase before public launch"],
]:
    ws5.append(item)
    r = ws5.max_row
    for cell in ws5[r]:
        cell.alignment = wrap
        cell.border = thin
    sev = item[1].upper()
    if sev in fills:
        ws5.cell(r, 2).fill = fills[sev]

# Column widths (Google Sheets compatible)
for ws in [ws1, ws2]:
    for col, w in [(1, 5), (2, 12), (3, 55), (4, 20), (5, 12), (6, 40)]:
        ws.column_dimensions[get_column_letter(col)].width = w

for col, w in [(1, 5), (2, 12), (3, 45), (4, 22), (5, 10), (6, 12), (7, 35)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

for ws in [ws4, ws5]:
    for col, w in [(1, 35), (2, 25), (3, 60)]:
        ws.column_dimensions[get_column_letter(col)].width = w

from openpyxl.utils import get_column_letter

out = r"C:\Users\ernes\Documents\Projects\WardPulse\comms\WardPulse_Action_Items.xlsx"
wb.save(out)
print(f"DONE: {out}")
